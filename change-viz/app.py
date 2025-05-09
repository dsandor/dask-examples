import pandas as pd
import json
import os
import numpy as np
from flask import Flask, render_template, request, jsonify, send_file
import tempfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.comments import Comment

app = Flask(__name__)

def load_data(csv_path, json_path):
    """Load CSV and changes JSON data"""
    try:
        # Load CSV file
        df = pd.read_csv(csv_path)
        
        # Load changes JSON file
        with open(json_path, 'r') as f:
            changes = json.load(f)
        
        return df, changes
    except Exception as e:
        return None, None, str(e)

def highlight_changes(df, changes):
    """Create a DataFrame with highlighted changes"""
    # Create a copy of the DataFrame to avoid modifying the original
    highlighted_df = df.copy()
    
    # Create a DataFrame to track changes for visualization
    changes_map = pd.DataFrame(index=df.index, columns=df.columns)
    changes_map = changes_map.fillna('')
    
    # ID column that contains the keys in the changes JSON
    id_column = 'ID_BB_GLOBAL'  # This might need to be adjusted based on the actual data
    
    # Process each change
    for id_value, change_data in changes.items():
        # Find rows with the matching ID
        matching_rows = df[df[id_column] == id_value].index.tolist()
        
        if not matching_rows:
            # Try alternative ID columns if the primary one doesn't match
            alternative_id_columns = ['ID_BB_SECURITY', 'ID_BB_YELLOWKEY', 'ID_BB_UNIQUE']
            for alt_col in alternative_id_columns:
                if alt_col in df.columns:
                    matching_rows = df[df[alt_col] == id_value].index.tolist()
                    if matching_rows:
                        break
        
        if matching_rows:
            for column_name, change_info in change_data.items():
                if column_name in df.columns:
                    for idx in matching_rows:
                        # Store the change information
                        changes_map.at[idx, column_name] = f"Previous: {change_info['previous_value']}"
    
    return highlighted_df, changes_map

def create_excel_with_highlights(df, changes_map, output_path):
    """Create an Excel file with highlighted changes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Changes Highlighted"
    
    # Convert DataFrame to rows and add to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Apply highlighting and comments for changed cells
            if r_idx > 1:  # Skip header row
                row_idx = r_idx - 2  # Adjust for 0-based indexing and header
                col_name = df.columns[c_idx - 1]
                
                if row_idx < len(changes_map) and col_name in changes_map.columns:
                    change_info = changes_map.iloc[row_idx][col_name]
                    if change_info:
                        # Highlight the cell
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        # Add comment with previous value
                        cell.comment = Comment(change_info, "Change Viz")
    
    # Save the workbook
    wb.save(output_path)
    return output_path

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    try:
        csv_path = request.form.get('csv_path', '')
        json_path = request.form.get('json_path', '')
        
        if not csv_path:
            csv_path = os.path.expanduser("~/Projects/custom_delta.csv")
        if not json_path:
            json_path = os.path.expanduser("~/Projects/custom_changes.json")
        
        # Load data
        df, changes = load_data(csv_path, json_path)
        
        if df is None:
            return jsonify({"error": f"Failed to load data: {changes}"})
        
        # Process changes
        highlighted_df, changes_map = highlight_changes(df, changes)
        
        # Create temporary file for Excel output
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        # Create Excel file with highlights
        output_path = create_excel_with_highlights(highlighted_df, changes_map, temp_file.name)
        
        # Return the path to the Excel file
        return jsonify({"success": True, "file_path": output_path})
    
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route('/download/<path:filename>')
def download_file(filename):
    return send_file(filename, as_attachment=True, download_name="changes_highlighted.xlsx")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
